class ReadWrite:
    def __init__(self):
        self.txtpath = r'C:\Users\Shi Yang\Desktop\test.txt'
        self.excelpath = r'C:\Users\Shi Yang\Desktop\学习\test.xlsx'
        self.yamlpath = r'C:\Users\Shi Yang\Desktop\users.yml'

    def txtread(self):
        list1 = []
        with open(self.txtpath,'r',encoding='utf-8') as f :
            values = f.readlines()
            f.close()
        for data in values:
                data_v = data.strip('\n')
                list1.append(data_v)
        return list1
    
    def txtwrite(self):
        username = input('输入用户名:')
        password = input('输入密码:')
        f = open(self.txtpath,'a',encoding='utf-8')
        values = f"{username},{password}\n"
        f.writelines(values) 
        f.close()

    def excelread(self,sheetname):
        import openpyxl
        wb = openpyxl.load_workbook(self.excelpath)
        table = wb[sheetname]
        rows = table.max_row
        cols = table.max_column
        list2 = []
        for row in range(2,rows+1):
            list1 = []
            for col in range(1,cols+1):
                values = table.cell(row,col).value
                list1.append(values)
            list2.append(list1)
        return list2

    def excelwrite(self,*args,sheetname): # *arg 元组
        import openpyxl
        wb = openpyxl.load_workbook(self.excelpath)
        table = wb[sheetname]
        rows = table.max_row
        cols = len(args)
        for col in range(cols):
            table.cell(rows+1,col+1).value = args[col]
        wb.save(self.excelpath)


    def mysqlread(self):
       import pymysql
       db = pymysql.connect(host = 'localhost',port = 3306,user = 'root',password = 'root',database = 'T32',charset = 'utf8')
       cur = db.cursor()
       sql = 'select * from users where username = "tester001"'
       cur.execute(sql)
       content = cur.fetchall() #返回元组
       return content


    def mysqlwrite(self,username,password):
        import pymysql
        db = pymysql.connect(host = 'localhost',port = 3306,user = 'root',password = 'root',database = 'T32',charset = 'utf8')
        cur = db.cursor()
        sql = f'insert into users values("{username}","{password}")'
        db.commit()

    def yamread(self):
        import yaml
        f = open(self.yamlpath,'r',encoding='utf-8')
        content = f.read()
        data = yaml.safe_load(content)  #转成字典格式
        return data
        
    def yamwrite(self,username,password):
        import yaml
        f = open(self.yamlpath,'a',encoding='utf-8')
        data = {'username':username,'password':password}
        yaml.safe_dump(data,f)
        f.close()
